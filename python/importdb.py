from openpyxl import load_workbook

import mysql.connector

mydb = mysql.connector.connect(
  host="localhost",
  user="root",
  password="",
  database="kki1"
)

mycursor = mydb.cursor(prepared=True)


wb = load_workbook(filename = 'sukajaya.xlsx')
sheet_ranges = wb['datawarga']
for x in sheet_ranges.iter_rows():
    kk = x[0].value
    nik = x[1].value
    nama = x[2].value
    jk = x[3].value
    tempat_l = x[4].value
    tgl_l = x[5].value
    gdr = x[6].value
    status = x[7].value
    agama = x[8].value
    pekerjaan = x[9].value
    dusun = x[10].value
    rt = x[11].value
    rw = x[12].value
    
    sql = "INSERT INTO masyarakat (nik, no_kk, nama, dusun, rt, rw, desa, kecamatan, kabupaten, tempat_lahir, tanggal_lahir, jenis_kelamin, agama, status_perkawinan, pekerjaan, gol_darah) VALUES (%s, %s,%s, %s, %s,%s,%s,%s,%s, %s, %s,%s,%s,%s,%s,%s)"
    val = (nik, kk, nama, dusun, rt, rw, "Sukajaya", "Rajadesa","Ciamis", tempat_l, tgl_l, jk, agama, status, pekerjaan, gdr)
    mycursor.execute(sql, val)


    sqlkk = "INSERT INTO kartu_keluarga (no_kk) VALUES (%s) ON DUPLICATE KEY UPDATE no_kk=%s"
    valkk = (kk,kk)
    mycursor.execute(sqlkk, valkk)
    mydb.commit()

    print(mycursor.rowcount, "record kk.")

# [   number:13.0,
#     text:'RAJADESA',
#     number:2007.0,
#     text:'SUKAJAYA',
#     text:'3207133110200005',
#     text:'3207135301510003',
#     text:'SUKANAH',
#     text:'Kepala Keluarga',
#     text:'Perempuan',
#     text:'CIAMIS',
#     text:'13-01-1951',
#     text:'-',
#     text:'Cerai Mati',
#     text:'Islam',
#     text:'Tamat SD/Sederajat',
#     text:'BURUH TANI/PERKEBUNAN',
#     text:'SARMINI',
#     text:'KARYA',
#     text:'CITAPEN PASIR',
#     text:'DUSUN CITAPEN PASIR',
#     number:12.0,
#     number:2.0,
#     number:46254.0
#     ]